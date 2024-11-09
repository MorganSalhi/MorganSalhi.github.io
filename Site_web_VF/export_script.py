import csv
import mysql.connector
import pandas as pd
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import sys
from datetime import datetime

# Database connection details
db_config = {
    'user': 'root',
    'password': 'root',
    'host': 'localhost',
    'database': 'thales11',
}

def fetch_bp_details_by_appart_and_assoc_ids(appart_ids, assoc_ids):
    """Fetch details of best practices based on appart and assoc IDs."""
    connection = mysql.connector.connect(**db_config)
    cursor = connection.cursor(dictionary=True)
    
    format_strings_appart = ','.join(['%s'] * len(appart_ids))
    format_strings_assoc = ','.join(['%s'] * len(assoc_ids))
    
    query = f"""
    SELECT 
        bonnespratiques.numBP, 
        bonnespratiques.texte AS Item, 
        GROUP_CONCAT(DISTINCT programmes.nomProgramme ORDER BY programmes.nomProgramme SEPARATOR ', ') AS Programme, 
        GROUP_CONCAT(DISTINCT phases.nomPhase ORDER BY phases.nomPhase SEPARATOR ', ') AS Phase, 
        GROUP_CONCAT(DISTINCT motscles.nomMotsCles ORDER BY motscles.nomMotsCles SEPARATOR ', ') AS `Mots clés`,
        '' AS Appliqué 
    FROM bonnespratiques
    LEFT JOIN appartenance ON bonnespratiques.numBP = appartenance.BP
    LEFT JOIN programmes ON programmes.numProg = appartenance.Programme
    LEFT JOIN phases ON phases.numPhases = appartenance.Phases
    LEFT JOIN association ON bonnespratiques.numBP = association.BP
    LEFT JOIN motscles ON association.numMC = motscles.numMC
    WHERE appartenance.numAppart IN ({format_strings_appart}) OR association.numAssoc IN ({format_strings_assoc})
    GROUP BY bonnespratiques.numBP
    """
    cursor.execute(query, tuple(appart_ids + assoc_ids))
    result = cursor.fetchall()
    
    cursor.close()
    connection.close()
    
    return result

def fetch_active_user():
    """Fetch the login of the currently active user from the database."""
    connection = mysql.connector.connect(**db_config)
    cursor = connection.cursor(dictionary=True)
    
    query = "SELECT login FROM utilisateurs WHERE statut = 'actif' LIMIT 1"
    cursor.execute(query)
    result = cursor.fetchone()
    
    cursor.close()
    connection.close()
    
    return result['login'] if result else 'None'

def export_to_excel(bp_details, user_id):
    """Export best practice details to an Excel file."""
    df = pd.DataFrame(bp_details)
    df = df[['Programme', 'Phase', 'Item', 'Mots clés', 'Appliqué']]
    
    file_path = "bp_details.xlsx"
    if os.path.exists(file_path):
        os.remove(file_path)
    
    wb = Workbook()
    ws = wb.active
    
    current_date = datetime.now().strftime("%d/%m/%Y")
    ws.append([f'Date: {current_date}', '', '', '', f'ID: {user_id}'])
    ws.append(['Programme', 'Phase', 'Item', 'Mots clés', 'Appliqué'])
    
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
    
    for cell in ws["2:2"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(file_path)
def export_to_pdf(bp_details, user_id):
    """Export best practice details to a PDF file."""
    file_path = "bp_details.pdf"
    if os.path.exists(file_path):
        os.remove(file_path)

    pdf = FPDF(orientation='L', format='A4')
    pdf.add_page()
    
    current_date = datetime.now().strftime("%d/%m/%Y")
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, txt=f"Date: {current_date}", ln=1, align="L")
    pdf.cell(0, 10, txt=f"ID: {user_id}", ln=1, align="L")
    
    # Column titles
    pdf.set_font("Arial", 'B', 12)
    col_widths = [50, 50, 90, 50, 30]
    pdf.cell(col_widths[0], 10, 'Programme', 1)
    pdf.cell(col_widths[1], 10, 'Phase', 1)
    pdf.cell(col_widths[2], 10, 'Item', 1)
    pdf.cell(col_widths[3], 10, 'Mots clés', 1)
    pdf.cell(col_widths[4], 10, 'Appliqué', 1)
    pdf.ln()
    
    # Data rows
    pdf.set_font("Arial", size=12)
    line_height = 10
    
    for bp in bp_details:
        x = pdf.get_x()
        y = pdf.get_y()
        
        # Draw each cell and get the height
        pdf.multi_cell(col_widths[0], line_height, bp['Programme'], 1)
        max_height = pdf.get_y() - y
        pdf.set_xy(x + col_widths[0], y)
        
        pdf.multi_cell(col_widths[1], line_height, bp['Phase'], 1)
        max_height = max(max_height, pdf.get_y() - y)
        pdf.set_xy(x + col_widths[0] + col_widths[1], y)
        
        pdf.multi_cell(col_widths[2], line_height, bp['Item'], 1)
        max_height = max(max_height, pdf.get_y() - y)
        pdf.set_xy(x + col_widths[0] + col_widths[1] + col_widths[2], y)
        
        pdf.multi_cell(col_widths[3], line_height, bp['Mots clés'], 1)
        max_height = max(max_height, pdf.get_y() - y)
        pdf.set_xy(x + col_widths[0] + col_widths[1] + col_widths[2] + col_widths[3], y)
        
        pdf.multi_cell(col_widths[4], line_height, bp['Appliqué'], 1)
        max_height = max(max_height, pdf.get_y() - y)
        
        pdf.ln(max_height)
    
    pdf.output(file_path)




def main(format, csv_file):
    appart_ids = []
    assoc_ids = []
    with open(csv_file, newline='') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)
        for row in reader:
            appart_ids.append(row[0])
            assoc_ids.append(row[1])
    
    bp_details = fetch_bp_details_by_appart_and_assoc_ids(appart_ids, assoc_ids)
    user_id = fetch_active_user()
    
    if format == 'excel':
        export_to_excel(bp_details, user_id)
    elif format == 'pdf':
        export_to_pdf(bp_details, user_id)
    else:
        print(f"Unknown format: {format}")

if __name__ == "__main__":
    format = sys.argv[1]
    csv_file = sys.argv[2]
    main(format, csv_file)
